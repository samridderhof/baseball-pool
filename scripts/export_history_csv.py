import csv
import sys
from pathlib import Path

from openpyxl import load_workbook


def main():
    if len(sys.argv) < 2:
        raise SystemExit(
            "Usage: python scripts/export_history_csv.py <input.xlsx> [output.csv]"
        )

    input_path = Path(sys.argv[1])
    output_path = (
        Path(sys.argv[2]) if len(sys.argv) > 2 else input_path.with_suffix(".history.csv")
    )

    workbook = load_workbook(input_path, data_only=True)
    sheet = workbook[workbook.sheetnames[0]]

    player_columns = []
    header_row = list(sheet.iter_rows(min_row=2, max_row=3, values_only=True))

    for column_index, player_name in enumerate(header_row[0], start=1):
        if not player_name or column_index <= 2:
            continue

        metric = header_row[1][column_index - 1]
        if metric == "CORRECT\nPICKS":
            player_columns.append(
                {
                    "player": str(player_name).strip(),
                    "correct_col": column_index,
                    "points_col": column_index + 1,
                    "cash_col": column_index + 3,
                }
            )

    with output_path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.writer(handle)
        writer.writerow(["week", "player", "correct_picks", "points", "cash_delta"])

        for row in sheet.iter_rows(min_row=4, values_only=True):
            week = row[0]
            if not isinstance(week, int):
                continue

            for player in player_columns:
                correct_picks = row[player["correct_col"] - 1]
                points = row[player["points_col"] - 1]
                cash_delta = row[player["cash_col"] - 1]

                if points is None:
                    continue

                writer.writerow(
                    [
                        week,
                        player["player"],
                        "" if correct_picks is None else int(correct_picks),
                        int(points),
                        int(cash_delta or 0),
                    ]
                )

    print(f"Wrote {output_path}")


if __name__ == "__main__":
    main()
